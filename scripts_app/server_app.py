#### Codigo desarrollado con asistencia de ChatGPT-4 de Open AI ####
import io
import os
import warnings
import hashlib
import pandas as pd
import flask
import base64
import time
import yaml
import shutil
# librerias para crear imagenes
import numpy as np
from PIL import Image
import pydicom
from uuid import uuid4
# librerías de dash
import dash_daq as daq
from dash import Dash, dcc, html, dash_table, ctx, no_update
import dash_bootstrap_components as dbc
from dash.exceptions import PreventUpdate
from dash.dependencies import Input, Output, State
# librerias de openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime, date, timedelta
# librerias de IA
from ultralytics import YOLO
warnings.filterwarnings("ignore")

#====================== Inicialización de la app ======================#
app = Dash(__name__, suppress_callback_exceptions=True,external_stylesheets=[dbc.themes.BOOTSTRAP])
server = app.server # necesario para almacenar sesiones en flask
server.secret_key = "my_secret_key_MAIA2025" #clave secreta
# Credenciales
VALID_USERNAME = "admin"
VALID_PASSWORD_HASH = hashlib.sha256("admin@123".encode()).hexdigest()  # Contraseña
# **Almacen de sesion en el navegador**
session_store = dcc.Store(id="session", storage_type="session")

# Ruta del directorio externo
@app.server.route("/external-images/<path:filename>")
def serve_external_images(filename):
    return flask.send_from_directory(EXTERNAL_IMAGES_PATH, filename)

# ruta del directorio de los archivos guardados temporalmente
@app.server.route("/external-images2/<path:filename>")
def serve_external_images2(filename):
    return flask.send_from_directory(save_path, filename)
#====================== Layout de la app ======================#
# layout de login"
login_layout = html.Div(
    [   
        # Capa semi trasparente
        html.Div(
            style={
                "position": "fixed",
                "top": "0",
                "left": "0",
                "width": "100%",
                "height": "100%",
                "backgroundColor": "rgba(0,0,0,0.5)",
                "zIndex": "-1",
            }
        ),
        # Tarjeta de login
        dbc.Card(
            dbc.CardBody(
                [
                    html.Img(src='/external-images/assets/logo.jpg', className="mb-3 d-block mx-auto", style={ "height": "80px"}),
                    html.H3("Cervical Clasify APP", className="text-center"),
                    dbc.Input(id="login-username", placeholder="Usuario", type="text", className="mb-3"),
                    dbc.Input(id="login-password", placeholder="Contraseña", type="password", className="mb-3"),
                    dbc.Button("Iniciar sesion", id="login-button", className="mb-3 w-100", style = {"backgroundColor": "#32CD32", "border": "none"}),
                    html.Div(id="login-message", className="text-danger text-center"),
                ]
            ),
            className="mx-auto mt-5",
            style={"width": "400px"},
        )
    ],
    style={"backgroundColor": "#f8f9fa", "height": "100vh", "display": "flex", "justifyContent": "center", "alignItems": "center"},
)
# Layout de la app principal
app_layout = html.Div(
    [
        # Barra de navegacion
        dbc.Navbar(
            dbc.Container(
                [
                    html.A(
                        dbc.Row(
                            [
                                dbc.Col(html.Img(src='/external-images/assets/logo.jpg', height="40px")),
                                dbc.Col(dbc.NavbarBrand("Cervical Clasify APP", className="ms-2")),
                            ],
                            align="center",
                            className="g-0",
                        ),
                        href="#",
                        style={"textDecoration": "none"},
                    ),
                    dbc.Button("Cerrar sesion", id="logout-button", color="danger"),
                ]
            ),
            color="light",
            dark=False,
        ),
        # Contenido principal
        html.Div(
            [
                html.H2("Bienvenido a la aplicacion de clasificacion de cancer cervical", className="text-center my-4"),
                html.P("Esta aplicacion permite cargar datos de pacientes y clasificarlos utilizando un modelo preentrenado.", className="text-center mb-4"),
                # Aqui puedes agregar m��s componentes para la funcionalidad principal
            ],
            className="container",
        ),
        dbc.Container([
            dbc.Card(
                [
                    dbc.CardHeader(html.H5("Carga y Procesamiento de Archivos")),
                    # Campo para cargar archivo .zip
                    dcc.Upload(
                        id='upload-data',
                        children=html.Div([
                            '📂 Arrastra y suelta o ',
                            html.A('Selecciona un archivo .zip con imagenes DICOM de corte axial T2')
                        ]),
                        style={
                            'width': '100%',
                            'height': '80px',
                            'lineHeight': '80px',
                            'borderWidth': '2px',
                            'borderStyle': 'dashed',
                            'borderRadius': '10px',
                            'textAlign': 'center',
                            'marginBottom': '15px',
                            'backgroundColor': '#f9f9f9',
                            'cursor': 'pointer'
                        },
                        multiple=False
                    ),
                    dbc.CardBody(
                        [
                            dbc.Button(
                                "Analizar archivos DICOM",
                                id="process-files-button",
                                color="primary",
                                className="w-100",
                                n_clicks=0
                            ),
                            html.Div(
                                [
                                    html.P("Progreso:", className="mt-3 mb-1"),
                                    dbc.Progress(id="progress-bar", value=0, striped=True, animated=True)
                                ],
                                style={"marginTop": "20px"}
                            ),
                            html.Div(id="loading-output", className="mt-3"),
                            html.Div(id='output-data-upload'),
                            dbc.Button(
                                "📥 Descargar Reporte Excel",
                                id="download-button",
                                color="success",
                                className="w-100 mt-3",
                                style={"display": "none"}  # oculto hasta que termine
                            ),
                            dcc.Download(id="download-excel"),
                            #html.A("Descargar", id="download-link", href="#", style={"display": "none"})
                        ]
                    )
                ],
                className="shadow-sm mb-4",
                style={"borderRadius": "12px"}
            ),
            # Intervalo para simular avance
            dcc.Interval(id="progress-interval", interval=800, disabled=True)  # 800 ms
        ]),
    ]
)
# # Layout principal
app.layout = html.Div(
    [
        dcc.Location(id="url", refresh=False),
        html.Div(id="page-content"),  # Contenedor donde se cargan los layouts
        session_store,  # Almacenar la sesi�n
    ]
)

########################## callbaks Manejo del Login ####################################
# Callback para manejar el login
@app.callback(
    Output("session", "data", allow_duplicate=True),
    Input("login-button", "n_clicks"),
    State("login-username", "value"),
    State("login-password", "value"),
    prevent_initial_call=True
)
def login(n_clicks, username, password):
    if username == VALID_USERNAME and hashlib.sha256(password.encode()).hexdigest() == VALID_PASSWORD_HASH:
        return {"logged_in": True}
    else:
        return {"logged_in": False}

# Callback para manejar el logout
@app.callback(
    Output("session", "data", allow_duplicate=True),
    Input("logout-button", "n_clicks"),
    prevent_initial_call=True
)
def logout(n_clicks):
    return {"logged_in": False}

# Callback para actualizar la interfaz al recargar
@app.callback(
    Output("page-content", "children"),
    Input("session", "data"),
    prevent_initial_call=True
)
def update_layout(session_data):
    if session_data and session_data.get("logged_in"):
        return app_layout
    return login_layout
########################## callbaks Manejo del aplicativo ####################################
# Callback para manejar la carga de archivos
@app.callback(
    Output('upload-data', 'children'),
    Input('upload-data', 'contents'),
    State('upload-data', 'filename'),
    prevent_initial_call=True
)
# Ejecutar la aplicación para recibir el archivo .zip, descomprimirlo y procesarlo contando los elementos en su interior
def handle_file_upload(contents, filename):
    if contents is None:
        raise PreventUpdate

    content_type, content_string = contents.split(',')
    decoded = io.BytesIO(base64.b64decode(content_string))

    # Aquí puedes agregar el código para descomprimir y procesar el archivo .zip
    # Por ejemplo, contar los archivos en su interior
    import zipfile
    with zipfile.ZipFile(decoded, 'r') as zip_ref:
        file_list = zip_ref.namelist()
        num_files = len(file_list)
        # llamar a la funcion que procesa los archivos y devuelve los resultados
        process_files(zip_ref)

    return f'Archivo "{filename}" cargado con éxito. Contiene {num_files} archivos.'
# Función para procesar los archivos descomprimidos, verificando que sean formato .mp3, los que coninicdan guardarlos en un path C:\Users\david\Downloads\archivos_app_guardados
def process_files(zip_ref):
    try:
        if not os.path.exists(save_path):
            os.makedirs(save_path)
        for file_name in zip_ref.namelist():
            if file_name.endswith('.dcm'):
                zip_ref.extract(file_name, save_path)
        print(f'Archivos .dcm guardados en {save_path}')
    except Exception as e:
        print(f'Error al procesar los archivos: {e}')
# Callback para procesar los archivos guardados al presionar el boton y actualiar barra de carga de 0 a 100% a medida que se procesan los archivos y previsualizar el resultado en una tabla con output-data-upload
progress_state = {"files": [], "current": 0, "total": 0, "results": []}


@app.callback(
    Output("progress-interval", "disabled"),
    Output("progress-bar", "value"),
    Output("progress-bar", "label"),
    Output("loading-output", "children"),
    Output("output-data-upload", "children"),
    Output("download-button", "style"),
    Input("process-files-button", "n_clicks"),
    Input("progress-interval", "n_intervals"),
    prevent_initial_call=True
)
def process_saved_files(n_clicks, n_intervals):
    global progress_state

    triggered_id = ctx.triggered_id

    # Paso 1: Botón presionado → inicializar progreso
    if triggered_id == "process-files-button":
        if not os.path.exists(save_path):
            return True, 0, "", "No hay archivos para procesar.", [], {"display": "none"}

        files = [f for f in os.listdir(save_path) if f.endswith(".dcm")]
        if not files:
            return True, 0, "", "No hay archivos .dcm para procesar.", [], {"display": "none"}

        progress_state = {"files": files, "current": 0, "total": len(files), "results": []}
        return False, 0, "0%", "Procesando archivos DICOM...", [], {"display": "none"}

    # Paso 2: Intervalo activo → procesar archivo por archivo
    if triggered_id == "progress-interval":
        if progress_state["current"] < progress_state["total"]:
            file_name = progress_state["files"][progress_state["current"]]

            # Procesar y clasificar
            result = process_dicom_and_classify(save_path, save_path, file_name)

            progress_state["results"].append(result)
            progress_state["current"] += 1

            percent = int(progress_state["current"] / progress_state["total"] * 100)
            return (
                False,  # seguir corriendo el Interval
                percent,
                f"{percent}%",
                f"Procesando {file_name}...",
                [],
                {"display": "none"}
            )
        else:
            # ✅ Terminó el procesamiento → añadir fila resumen
            from collections import Counter

            clases = [r["Clasificación"] for r in progress_state["results"]]
            conteo = Counter(clases)

            if conteo:
                clase_mayoritaria, cantidad = conteo.most_common(1)[0]
                total = len(clases)
                porcentaje = (cantidad / total) * 100 if total > 0 else 0

                resumen = {
                    "Archivo DICOM": "Confianza en Paciente",
                    "Clasificación": clase_mayoritaria,
                    "% Probabilidad": f"{porcentaje:.2f}%",
                    "Imagen": ""
                }
                progress_state["results"].append(resumen)

            # Construir tabla
            table = dash_table.DataTable(
                columns=[
                    {"name": "Archivo DICOM", "id": "Archivo DICOM"},
                    {"name": "Clasificación", "id": "Clasificación"},
                    {"name": "% Probabilidad", "id": "% Probabilidad"},
                    {"name": "Imagen", "id": "Imagen", "presentation": "markdown"},
                ],
                data=progress_state["results"],
                style_table={"overflowX": "auto"},
                style_cell={"textAlign": "center"},
            )

            return True, 100, "100%", "Procesamiento completado ✅", table, {"display": "block"}

    raise PreventUpdate


# Función para convertir DICOM a PNG
def process_dicom_and_classify(input_path, output_path, filename, size=(224, 224)):
    """
    Convierte un DICOM a PNG y lo clasifica con YOLO.
    Retorna clasificación y % confianza.
    """
    file_path = os.path.join(input_path, filename)
    ds = pydicom.dcmread(file_path)

    # Extraer el pixel array
    img = ds.pixel_array.astype(float)
    img = (np.maximum(img, 0) / img.max()) * 255.0
    img = np.uint8(img)

    # Convertir a imagen PIL
    im = Image.fromarray(img).convert("L")
    im = im.resize(size)
    im = im.convert("RGB")

    os.makedirs(output_path, exist_ok=True)
    out_file = os.path.splitext(filename)[0] + ".png"
    out_path = os.path.join(output_path, out_file)
    im.save(out_path)

    # 🔥 Clasificación con YOLO
    results = model(out_path)  # correr modelo sobre la imagen PNG
    if results and len(results[0].probs) > 0:
        # Obtener clase con mayor probabilidad
        class_id = int(results[0].probs.top1)
        class_name = results[0].names[class_id]
        confidence = float(results[0].probs.top1conf) * 100
        if "class_0" in class_name:
            class_name = "Cervical Squamous Cell Carcinoma"
        elif "class_1" in class_name:
            class_name = "Mucinous Adenocarcinoma of Endocervical Type"
    else:
        class_name = "Desconocido"
        confidence = 0.0
    # Crear URL servida por Flask
    url = f"/external-images2/{out_file}?v={uuid4()}"
    return {
        "Archivo DICOM": filename,
        "Clasificación": class_name,
        "% Probabilidad": f"{confidence:.2f}%",
        "Imagen": f"![png]({url})"
    }

@app.callback(
    Output("download-excel", "data"),
    Input("download-button", "n_clicks"),
    prevent_initial_call=True
)
def download_excel(n_clicks):
    if n_clicks is None:
        raise PreventUpdate
    # ✅ Terminó el procesamiento
    # Vaciar carpeta save_path
    for f in os.listdir(save_path):
        file_path = os.path.join(save_path, f)
        try:
            if os.path.isfile(file_path):
                os.remove(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)
        except Exception as e:
            print(f"Error eliminando {file_path}: {e}")
    try:
        # Convertimos los resultados a DataFrame y filtramos columnas relevantes
        df = pd.DataFrame(progress_state["results"])[
            ["Archivo DICOM", "Clasificación", "% Confianza"]
        ]


        # Crear un archivo Excel en memoria con openpyxl
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Resultados"

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        # (Opcional) estilizar, insertar imágenes, etc.
        ws["A1"].style = "Title"

        # Guardar en memoria
        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)

        # Retornar el archivo como descarga
        return dcc.send_bytes(stream.getvalue(), "resultados.xlsx")
    except Exception as e:
        print(f"Error al generar el archivo Excel: {e}")
        raise PreventUpdate

# funcion para cargar datos del archivo config.yaml
def load_config(config_path='config/config.yaml'):
    with open(config_path, 'r') as file:
        config = yaml.safe_load(file)
    return config

if __name__ == '__main__':
    # declarar como global las variables de path
    global EXTERNAL_IMAGES_PATH, save_path, model
    # Cargar configuracion
    config = load_config()
    EXTERNAL_IMAGES_PATH = config['paths']['EXTERNAL_IMAGES_PATH']
    save_path = config['paths']['save_path']
    model_path = config['paths']['model_path']
    # Cargar modelo
    model = YOLO(model_path)
    # Ejecutar la app
    app.run_server(host='0.0.0.0', port=8050, debug=True)